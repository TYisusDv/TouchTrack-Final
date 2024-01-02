
from flask import Flask, render_template, request, redirect, url_for, session, make_response, send_file, jsonify, send_from_directory
from datetime import datetime, timedelta, timezone
from passlib.hash import bcrypt
from urllib.parse import urlencode, urlparse, urlunparse, parse_qsl
from PIL import Image
from markupsafe import Markup
from openpyxl.styles import PatternFill
import json, uuid, requests, re, math, time, sys, random, shutil, os, base64, subprocess, psutil, glob, socket, string, html, secrets, hashlib, jwt, pytz, openpyxl, mimetypes

config_hostname = socket.gethostname()

config_app = {
    'debug': True,
    'db_mongo': {
        'main': {
            'name': 'touchtrack'
        }
    },
    'url_main': 'http://127.0.0.1:5000',
}

config_routes_nocache = [
    '/auth/sign-in',
    '/auth/sign-up',
    '/auth/logout'
]

config_routes = [
    '',
    'dashboard',
    'manage/users',
    'manage/user/add',
    'manage/user/edit',
    'manage/students',
    'manage/student/add',
    'manage/student/edit',
    'manage/student/fingerprints',
    'manage/groups',
    'manage/group/add',
    'manage/group/edit',
    'manage/careers',
    'manage/career/add',
    'manage/career/edit',
    'manage/events',
    'manage/event/add',
    'manage/event/edit',
    'manage/event/attendances',
    'manage/attendances',
]

def config_splitList(s, text, n):
    try:
        return text.split(s)[n] if len(text.split(s)) > n else None
    except:
        return None

def config_validateForm(empty = False, form = None, min = None, max = None):
    if not empty:
        if not form:
            return False

        if not min is None:
            if len(form) < min:
                return False
            
        if not max is None:
            if len(form) > max:
                return False
                
    return True

def config_genUniqueID():
    datetime_now = datetime.now()
    salt = str(datetime_now.timestamp())
    token = secrets.token_hex(9)
    unique_string = hashlib.sha256((salt + token).encode('utf-8')).hexdigest()[:18]
    return unique_string.upper()

def config_verifyText(text):
    return all(char.isalpha() or char.isspace() for char in text)

def config_isValidURL(url):
    parsed_url = urlparse(url)
    return parsed_url.scheme and parsed_url.netloc

def config_urlParam(url, param, value):
    parsed_url = urlparse(url)
    query_params = parse_qsl(parsed_url.query)
    query_params.append((param, value))
    new_query = urlencode(query_params)
    new_url = urlunparse(
        (parsed_url.scheme, parsed_url.netloc, parsed_url.path, parsed_url.params, new_query, parsed_url.fragment)
    )
    return new_url

def config_searchRegex(search = None):
    if search:
        words = search.split()
        query_regex = ".*" + ".*".join(words) + ".*"
        return query_regex
    else:
        return ''

def config_convertDate(date):
    date_gmt = datetime.strptime(str(date), '%Y-%m-%d %H:%M:%S.%f')
    date_gmt = pytz.timezone('GMT').localize(date_gmt)

    newdate = date_gmt.astimezone(pytz.timezone('America/Mexico_City'))

    date_converted = newdate.strftime('%d/%m/%Y %I:%M %p')
    
    return date_converted

def config_convertLocalDate(date):
    date = datetime.strptime(str(date), '%Y-%m-%d %H:%M:%S')    
    date_gmt = pytz.timezone('GMT').localize(date)
    newdate = date_gmt.astimezone(pytz.timezone('America/Mexico_City'))
    date_converted = newdate.strftime('%d/%m/%Y %I:%M %p')    
    return date_converted

def config_convertStringDate(date):
    date_object = datetime.fromisoformat(date)    
    return date_object 

def config_convertStringDate2(date):
    date_object = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    return date_object

def config_convertDatetoT(date):
    date = datetime.strptime(str(date), '%Y-%m-%d %H:%M:%S')    
    date_gmt = pytz.timezone('GMT').localize(date)
    newdate = date_gmt.astimezone(pytz.timezone('America/Mexico_City'))
    date_converted = newdate.strftime('%Y-%m-%dT%H:%M')    
    return date_converted

def config_desktop_msg(msg):
    return f'<html><p>{msg}</p></html>'

def config_allowedFile(filename):
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def config_create_worksheet(workbook, title, headers):
    worksheet = workbook.create_sheet(title=title)
    for i, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=i, value=header)
    return worksheet